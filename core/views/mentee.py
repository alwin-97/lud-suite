import csv
import openpyxl

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from core.decorators import role_required
from core.forms import ObjectiveItemForm, YearPlanItemForm
from core.models import Mentee, MenteeAssessment, ObjectiveItem, RatingDomain, YearPlanItem
from core.views.common import get_mentee_for_user, mentor_can_access_mentee


MENTEE_DASHBOARD_TEMPLATE = "core/mentee/dashboard.html"
MENTEE_OBJECTIVE_LIST_TEMPLATE = "core/mentee/objectives/list.html"
MENTEE_OBJECTIVE_FORM_TEMPLATE = "core/mentee/objectives/form.html"
MENTEE_YEAR_PLAN_LIST_TEMPLATE = "core/mentee/year_plans/list.html"
MENTEE_YEAR_PLAN_FORM_TEMPLATE = "core/mentee/year_plans/form.html"
MENTEE_ASSESSMENT_LIST_TEMPLATE = "core/mentee/assessments/list.html"


@role_required(allowed_roles=["mentee"])
def mentee_dashboard_view(request):
    mentee = get_mentee_for_user(request.user)
    if not mentee:
        messages.warning(request, "Your mentee profile is not set up yet.")
        return render(request, MENTEE_DASHBOARD_TEMPLATE, {"mentee": None, "active_page": "dashboard"})

    objectives = ObjectiveItem.objects.filter(mentee=mentee)
    year_plans = YearPlanItem.objects.filter(mentee=mentee)

    objective_status_counts = (
        objectives.values("status__name")
        .annotate(count=Count("id"))
        .order_by("status__name")
    )
    year_plan_counts = year_plans.values("year").annotate(count=Count("id")).order_by("year")

    context = {
        "mentee": mentee,
        "objectives": objectives[:5],
        "year_plans": year_plans[:5],
        "objective_status_counts": objective_status_counts,
        "year_plan_counts": year_plan_counts,
        "active_page": "dashboard",
    }
    return render(request, MENTEE_DASHBOARD_TEMPLATE, context)


@login_required
@role_required(allowed_roles=["mentee"])
def objective_list_view(request):
    mentee = get_mentee_for_user(request.user)
    if not mentee:
        messages.warning(request, "Your mentee profile is not set up yet.")
        return render(request, MENTEE_OBJECTIVE_LIST_TEMPLATE, {"mentee": None, "active_page": "objectives"})
    objectives = ObjectiveItem.objects.filter(mentee=mentee).select_related("status")
    return render(request, MENTEE_OBJECTIVE_LIST_TEMPLATE, {"mentee": mentee, "objectives": objectives, "active_page": "objectives"})


@login_required
@role_required(allowed_roles=["mentee"])
def objective_create_view(request):
    mentee = get_mentee_for_user(request.user)
    if not mentee:
        messages.warning(request, "Your mentee profile is not set up yet.")
        return redirect("mentee_dashboard")
    if request.method == "POST":
        form = ObjectiveItemForm(request.POST, request.FILES)
        if form.is_valid():
            objective = form.save(commit=False)
            objective.mentee = mentee
            objective.created_by = request.user
            objective.updated_by = request.user
            objective.save()
            messages.success(request, "Objective saved.")
            return redirect("objective_list")
    else:
        form = ObjectiveItemForm()
    return render(request, MENTEE_OBJECTIVE_FORM_TEMPLATE, {"form": form, "mode": "create", "active_page": "objectives"})


@login_required
@role_required(allowed_roles=["mentee"])
def objective_edit_view(request, objective_id):
    mentee = get_mentee_for_user(request.user)
    if not mentee:
        messages.warning(request, "Your mentee profile is not set up yet.")
        return redirect("mentee_dashboard")
    objective = get_object_or_404(ObjectiveItem, id=objective_id, mentee=mentee)
    if request.method == "POST":
        form = ObjectiveItemForm(request.POST, request.FILES, instance=objective)
        if form.is_valid():
            objective = form.save(commit=False)
            objective.updated_by = request.user
            objective.save()
            messages.success(request, "Objective updated.")
            return redirect("objective_list")
    else:
        form = ObjectiveItemForm(instance=objective)
    return render(
        request,
        MENTEE_OBJECTIVE_FORM_TEMPLATE,
        {"form": form, "mode": "edit", "objective": objective, "active_page": "objectives"},
    )


@login_required
@role_required(allowed_roles=["mentee"])
def year_plan_list_view(request):
    mentee = get_mentee_for_user(request.user)
    if not mentee:
        messages.warning(request, "Your mentee profile is not set up yet.")
        return render(request, MENTEE_YEAR_PLAN_LIST_TEMPLATE, {"mentee": None, "active_page": "year_plans"})
    year_plans = YearPlanItem.objects.filter(mentee=mentee).select_related("status")
    return render(request, MENTEE_YEAR_PLAN_LIST_TEMPLATE, {"mentee": mentee, "year_plans": year_plans, "active_page": "year_plans"})


@login_required
@role_required(allowed_roles=["mentee"])
def year_plan_create_view(request):
    mentee = get_mentee_for_user(request.user)
    if not mentee:
        messages.warning(request, "Your mentee profile is not set up yet.")
        return redirect("mentee_dashboard")
    if request.method == "POST":
        form = YearPlanItemForm(request.POST)
        if form.is_valid():
            year_plan = form.save(commit=False)
            year_plan.mentee = mentee
            year_plan.created_by = request.user
            year_plan.updated_by = request.user
            year_plan.save()
            messages.success(request, "Year plan saved.")
            return redirect("year_plan_list")
    else:
        form = YearPlanItemForm()
    return render(request, MENTEE_YEAR_PLAN_FORM_TEMPLATE, {"form": form, "mode": "create", "active_page": "year_plans"})


@login_required
@role_required(allowed_roles=["mentee"])
def year_plan_edit_view(request, year_plan_id):
    mentee = get_mentee_for_user(request.user)
    if not mentee:
        messages.warning(request, "Your mentee profile is not set up yet.")
        return redirect("mentee_dashboard")
    year_plan = get_object_or_404(YearPlanItem, id=year_plan_id, mentee=mentee)
    if request.method == "POST":
        form = YearPlanItemForm(request.POST, instance=year_plan)
        if form.is_valid():
            year_plan = form.save(commit=False)
            year_plan.updated_by = request.user
            year_plan.save()
            messages.success(request, "Year plan updated.")
            return redirect("year_plan_list")
    else:
        form = YearPlanItemForm(instance=year_plan)
    return render(
        request,
        MENTEE_YEAR_PLAN_FORM_TEMPLATE,
        {"form": form, "mode": "edit", "year_plan": year_plan, "active_page": "year_plans"},
    )


@login_required
@role_required(allowed_roles=["mentee"])
def mentee_assessment_list_view(request):
    mentee = get_mentee_for_user(request.user)
    if not mentee:
        messages.warning(request, "Your mentee profile is not set up yet.")
        return render(request, MENTEE_ASSESSMENT_LIST_TEMPLATE, {"mentee": None, "active_page": "assessments"})
    assessments = MenteeAssessment.objects.filter(mentee=mentee).select_related("session_type")
    return render(
        request,
        MENTEE_ASSESSMENT_LIST_TEMPLATE,
        {"mentee": mentee, "assessments": assessments, "active_page": "assessments"},
    )


@login_required
def export_mentee_progress_csv(request, mentee_id):
    mentee = get_object_or_404(Mentee, id=mentee_id)
    if request.user.role == "mentee" and mentee.user_id != request.user.id:
        return HttpResponse("You are not allowed to export this mentee.", status=403)
    if request.user.role == "mentor" and not mentor_can_access_mentee(request.user, mentee):
        return HttpResponse("You are not allowed to export this mentee.", status=403)

    objectives = ObjectiveItem.objects.filter(mentee=mentee).select_related("status")
    year_plans = YearPlanItem.objects.filter(mentee=mentee).select_related("status")
    assessments = (
        MenteeAssessment.objects.filter(mentee=mentee)
        .select_related("session_type")
        .prefetch_related("ratings__domain")
    )
    assessment_domains = list(
        RatingDomain.objects.filter(assessmentrating__assessment__mentee=mentee)
        .distinct()
        .order_by("year", "sort_order", "name")
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{mentee.id}_progress.csv"'
    writer = csv.writer(response)

    writer.writerow(["Objectives"])
    writer.writerow(
        [
            "Title",
            "Objective Text",
            "Action Items",
            "Start Date",
            "End Date",
            "Expected Outcome",
            "Status",
            "Progress %",
            "Mentee Remarks",
            "Mentor Comments",
        ]
    )
    for obj in objectives:
        writer.writerow(
            [
                obj.objective_title,
                obj.objective_text,
                obj.action_items,
                obj.start_date,
                obj.end_date,
                obj.expected_outcome,
                obj.status.name if obj.status else "",
                obj.progress_percent,
                obj.mentee_remarks,
                obj.mentor_comments,
            ]
        )

    writer.writerow([])
    writer.writerow(["Year Plans"])
    writer.writerow(
        [
            "Year",
            "Milestone",
            "Deliverable",
            "Target Date",
            "Target Period",
            "Status",
            "Remarks",
            "Mentor Comments",
            "Review Date",
        ]
    )
    for plan in year_plans:
        writer.writerow(
            [
                plan.get_year_display(),
                plan.milestone,
                plan.deliverable,
                plan.target_date,
                plan.target_period,
                plan.status.name if plan.status else "",
                plan.remarks,
                plan.mentor_comments,
                plan.review_date,
            ]
        )

    writer.writerow([])
    writer.writerow(["Assessments"])
    writer.writerow(
        [
            "Date",
            "Year",
            "Session Type",
            "Theme/Topic",
            "Beginning Mood",
            "End Mood",
            "Average",
            "Mentor Remarks",
            "Action Plan",
            *[f"{domain.get_year_display()} - {domain.name}" for domain in assessment_domains],
        ]
    )
    for assessment in assessments:
        rating_map = {rating.domain_id: rating.value for rating in assessment.ratings.all()}
        writer.writerow(
            [
                assessment.date,
                assessment.get_year_display(),
                assessment.session_type,
                assessment.theme_topic,
                assessment.beginning_mood,
                assessment.end_mood,
                assessment.average_score(),
                assessment.mentor_remarks,
                assessment.action_plan,
                *[rating_map.get(domain.id, "") for domain in assessment_domains],
            ]
        )

    return response


@login_required
def export_mentee_progress_excel(request, mentee_id):
    mentee = get_object_or_404(Mentee, id=mentee_id)
    if request.user.role == "mentee" and mentee.user_id != request.user.id:
        return HttpResponse("You are not allowed to export this mentee.", status=403)
    if request.user.role == "mentor" and not mentor_can_access_mentee(request.user, mentee):
        return HttpResponse("You are not allowed to export this mentee.", status=403)

    objectives = ObjectiveItem.objects.filter(mentee=mentee).select_related("status")
    year_plans = YearPlanItem.objects.filter(mentee=mentee).select_related("status")
    assessments = (
        MenteeAssessment.objects.filter(mentee=mentee)
        .select_related("session_type")
        .prefetch_related("ratings__domain")
    )
    assessment_domains = list(
        RatingDomain.objects.filter(assessmentrating__assessment__mentee=mentee)
        .distinct()
        .order_by("year", "sort_order", "name")
    )

    workbook = openpyxl.Workbook()
    
    # 1. Objectives Sheet
    ws_obj = workbook.active
    ws_obj.title = "Objectives"
    ws_obj.append(["Title", "Objective Text", "Action Items", "Start Date", "End Date", "Expected Outcome", "Status", "Progress %", "Mentee Remarks", "Mentor Comments"])
    for obj in objectives:
        ws_obj.append([
            obj.objective_title, obj.objective_text, obj.action_items, 
            obj.start_date.strftime("%Y-%m-%d") if obj.start_date else "", 
            obj.end_date.strftime("%Y-%m-%d") if obj.end_date else "", 
            obj.expected_outcome, obj.status.name if obj.status else "", 
            obj.progress_percent, obj.mentee_remarks, obj.mentor_comments
        ])
        
    # 2. Year Plans Sheet
    ws_year = workbook.create_sheet("Year Plans")
    ws_year.append(["Year", "Milestone", "Deliverable", "Target Date", "Target Period", "Status", "Remarks", "Mentor Comments", "Review Date"])
    for plan in year_plans:
        ws_year.append([
            plan.get_year_display(), plan.milestone, plan.deliverable, 
            plan.target_date.strftime("%Y-%m-%d") if plan.target_date else "", 
            plan.target_period, plan.status.name if plan.status else "", 
            plan.remarks, plan.mentor_comments, 
            plan.review_date.strftime("%Y-%m-%d") if plan.review_date else ""
        ])
        
    # 3. Assessments Sheet
    ws_assess = workbook.create_sheet("Assessments")
    domain_headers = [f"{domain.get_year_display()} - {domain.name}" for domain in assessment_domains]
    ws_assess.append(["Date", "Year", "Session Type", "Theme/Topic", "Beginning Mood", "End Mood", "Average", "Mentor Remarks", "Action Plan"] + domain_headers)
    
    for assessment in assessments:
        rating_map = {rating.domain_id: rating.value for rating in assessment.ratings.all()}
        row = [
            assessment.date.strftime("%Y-%m-%d") if assessment.date else "",
            assessment.get_year_display(),
            assessment.session_type.name if assessment.session_type else "",
            assessment.theme_topic,
            assessment.beginning_mood.name if assessment.beginning_mood else "",
            assessment.end_mood.name if assessment.end_mood else "",
            assessment.average_score(),
            assessment.mentor_remarks,
            assessment.action_plan
        ]
        row.extend([rating_map.get(domain.id, "") for domain in assessment_domains])
        ws_assess.append(row)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{mentee.id}_progress.xlsx"'
    workbook.save(response)
    return response
