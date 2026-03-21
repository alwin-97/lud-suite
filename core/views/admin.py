import csv
import traceback

import openpyxl
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.views.decorators.http import require_POST

from core.decorators import role_required
from core.forms import (
    AcademicCycleForm,
    ChapterForm,
    DomainIndicatorForm,
    LocationForm,
    MentorMenteeAssignmentForm,
    MoodCategoryForm,
    NotificationForm,
    ProgrammeForm,
    RatingDomainForm,
    RatingScaleDefinitionForm,
    ReferenceContentForm,
    SchoolForm,
    SessionTypeForm,
    StatusConfigForm,
    TemplateConfigForm,
    VolunteerReportingAssignmentForm,
)
from core.models import (
    AcademicCycle,
    Activity,
    Chapter,
    CustomUser,
    DomainIndicator,
    Location,
    Mentee,
    MenteeAssessment,
    MenteeUploadLog,
    MentorMenteeAssignment,
    MoodCategory,
    Notification,
    Programme,
    RatingDomain,
    RatingScaleDefinition,
    ReferenceContent,
    School,
    SessionType,
    StatusConfig,
    TemplateConfig,
    DiaryEntry,
    ReflectiveReport,
    RepositoryAsset,
    VolunteerTranscript,
    VolunteerReportingAssignment,
)
from core.roles import ROLE_CHOICES, ROLE_KEYS
from core.views.common import User


ADMIN_DASHBOARD_TEMPLATE = "core/admin/dashboard.html"
ADMIN_CONFIG_HOME_TEMPLATE = "core/admin/config/home.html"
ADMIN_CONFIG_LIST_TEMPLATE = "core/admin/config/list.html"
ADMIN_CONFIG_FORM_TEMPLATE = "core/admin/config/form.html"
ADMIN_CONFIG_DELETE_TEMPLATE = "core/admin/config/delete.html"
ADMIN_USERS_MANAGE_TEMPLATE = "core/admin/users/manage.html"
ADMIN_USERS_ROLE_MANAGER_TEMPLATE = "core/admin/users/role_manager.html"
ADMIN_USERS_DETAIL_TEMPLATE = "core/admin/users/detail.html"
ADMIN_ASSIGNMENTS_MENTOR_TEMPLATE = "core/admin/assignments/manage_endorser_mentors.html"
ADMIN_ASSIGNMENTS_MENTEE_TEMPLATE = "core/admin/assignments/manage_mentee_assignments.html"
ADMIN_ASSIGNMENTS_VOLUNTEER_REPORTING_TEMPLATE = "core/admin/assignments/manage_volunteer_reporting_assignments.html"
ADMIN_NOTIFICATIONS_TEMPLATE = "core/admin/notifications/manage.html"
ADMIN_MENTEE_UPLOAD_TEMPLATE = "core/admin/mentee_upload/upload.html"
ADMIN_MENTEE_UPLOAD_LOG_TEMPLATE = "core/admin/mentee_upload/log.html"


CONFIG_REGISTRY = {
    "schools": {"model": School, "form": SchoolForm, "title": "Schools", "icon": "fa-school"},
    "locations": {"model": Location, "form": LocationForm, "title": "Locations", "icon": "fa-location-dot"},
    "chapters": {"model": Chapter, "form": ChapterForm, "title": "Chapters", "icon": "fa-bookmark"},
    "academic-cycles": {"model": AcademicCycle, "form": AcademicCycleForm, "title": "Academic Cycles", "icon": "fa-calendar-alt"},
    "programmes": {"model": Programme, "form": ProgrammeForm, "title": "Programmes", "icon": "fa-graduation-cap"},
    "statuses": {"model": StatusConfig, "form": StatusConfigForm, "title": "Statuses", "icon": "fa-tags"},
    "session-types": {"model": SessionType, "form": SessionTypeForm, "title": "Session Types", "icon": "fa-calendar-days"},
    "rating-domains": {"model": RatingDomain, "form": RatingDomainForm, "title": "Rating Domains", "icon": "fa-chart-line"},
    "domain-indicators": {"model": DomainIndicator, "form": DomainIndicatorForm, "title": "Domain Indicators", "icon": "fa-list-check"},
    "rating-scale-definitions": {
        "model": RatingScaleDefinition,
        "form": RatingScaleDefinitionForm,
        "title": "Rating Scales",
        "icon": "fa-ranking-star",
    },
    "mood-categories": {"model": MoodCategory, "form": MoodCategoryForm, "title": "Mood Categories", "icon": "fa-face-smile"},
    "reference-content": {"model": ReferenceContent, "form": ReferenceContentForm, "title": "Reference Content", "icon": "fa-book-open"},
    "template-configs": {"model": TemplateConfig, "form": TemplateConfigForm, "title": "Template Configs", "icon": "fa-sliders"},
}


@role_required(allowed_roles=["admin"])
def edit_user_view(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.method != "POST":
        return redirect("manage_user")

    first_name = request.POST.get("first_name")
    last_name = request.POST.get("last_name")
    email = request.POST.get("email")
    phone = request.POST.get("phone")
    role = request.POST.get("role")
    password = request.POST.get("password")
    current_year = request.POST.get("current_year")

    if not (first_name and last_name and email and phone and role):
        messages.error(request, "All fields are required.")
        return redirect("manage_user")
    if role not in ROLE_KEYS:
        messages.error(request, "Invalid role selected.")
        return redirect("manage_user")

    if User.objects.filter(email=email).exclude(id=user.id).exists():
        messages.error(request, "User with this email already exists.")
        return redirect("manage_user")

    user.first_name = first_name
    user.last_name = last_name
    user.email = email
    user.username = email
    user.phone = phone
    user.role = role
    roles = [item for item in request.POST.getlist("roles") if item in ROLE_KEYS]
    if roles:
        if role not in roles:
            roles.insert(0, role)
        user.roles = roles
    else:
        user.roles = [role]
    if password:
        user.set_password(password)
    user.save()

    if role == "mentee":
        mentee, created = Mentee.objects.get_or_create(
            user=user,
            defaults={"full_name": f"{first_name} {last_name}".strip(), "current_year": int(current_year) if current_year else 1},
        )
        if not created:
            mentee.full_name = f"{first_name} {last_name}".strip()
            if current_year:
                mentee.current_year = int(current_year)
            mentee.save()

    messages.success(request, "User updated successfully.")
    return redirect("manage_user")


@role_required(allowed_roles=["admin"])
def delete_user_view(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, "User deleted successfully.")
    return redirect("manage_user")


@role_required(allowed_roles=["admin"])
def admin_dashboard_view(request):
    total_mentors = CustomUser.objects.filter(role="mentor").count()
    total_endorsers = CustomUser.objects.filter(role="endorser").count()
    total_mentees = Mentee.objects.filter(is_active=True).count()
    total_assignments = MentorMenteeAssignment.objects.count()
    total_schools = School.objects.count()
    total_chapters = Chapter.objects.filter(is_active=True).count()
    total_locations = Location.objects.filter(is_active=True).count()
    total_programmes = Programme.objects.filter(is_active=True).count()
    active_cycle = AcademicCycle.objects.filter(is_active=True).first()
    total_assessments = MenteeAssessment.objects.count()
    total_users = CustomUser.objects.exclude(role="admin").count()
    total_volunteers = CustomUser.objects.filter(role="volunteer").count()
    volunteer_reporting_assignments = VolunteerReportingAssignment.objects.count()
    approved_reports = ReflectiveReport.objects.filter(status="Approved").count()
    approved_transcripts = VolunteerTranscript.objects.filter(approval_status="Approved").count()
    repository_assets = RepositoryAsset.objects.count()
    assigned_endorsers = CustomUser.objects.filter(role="endorser", mentors__isnull=False).distinct().count()
    assigned_mentors = CustomUser.objects.filter(role="mentor", assigned_endorsers__isnull=False).distinct().count()
    recent_notifications = Notification.objects.order_by("-created_at")[:5]
    recent_uploads = MenteeUploadLog.objects.order_by("-created_at")[:5]
    context = {
        "total_mentors": total_mentors,
        "total_endorsers": total_endorsers,
        "total_mentees": total_mentees,
        "total_assignments": total_assignments,
        "total_schools": total_schools,
        "total_chapters": total_chapters,
        "total_locations": total_locations,
        "total_programmes": total_programmes,
        "active_cycle": active_cycle,
        "total_assessments": total_assessments,
        "total_users": total_users,
        "total_volunteers": total_volunteers,
        "volunteer_reporting_assignments": volunteer_reporting_assignments,
        "unmapped_volunteers": max(total_volunteers - volunteer_reporting_assignments, 0),
        "approved_reports": approved_reports,
        "approved_transcripts": approved_transcripts,
        "repository_assets": repository_assets,
        "assigned_endorsers": assigned_endorsers,
        "unassigned_endorsers": total_endorsers - assigned_endorsers,
        "assigned_mentors": assigned_mentors,
        "unassigned_mentors": total_mentors - assigned_mentors,
        "recent_notifications": recent_notifications,
        "recent_uploads": recent_uploads,
        "active_page": "dashboard",
    }
    return render(request, ADMIN_DASHBOARD_TEMPLATE, context)


@role_required(allowed_roles=["admin"])
def admin_config_home_view(request):
    config_items = [
        {"key": key, "label": val["title"], "icon": val.get("icon", "fa-gear")}
        for key, val in CONFIG_REGISTRY.items()
    ]
    return render(request, ADMIN_CONFIG_HOME_TEMPLATE, {"config_items": config_items, "active_page": "config"})


@role_required(allowed_roles=["admin"])
def admin_config_list_view(request, config_key):
    config = CONFIG_REGISTRY.get(config_key)
    if not config:
        return HttpResponse("Config not found.", status=404)

    model_class = config["model"]
    form_class = config["form"]
    title = config["title"]
    items = model_class.objects.all()

    if request.method == "POST":
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f"{title} saved.")
            return redirect("admin_config_list", config_key=config_key)
    else:
        form = form_class()

    items_with_forms = [(item, form_class(instance=item)) for item in items]

    return render(
        request,
        ADMIN_CONFIG_LIST_TEMPLATE,
        {"items": items, "items_with_forms": items_with_forms, "form": form, "title": title, "config_key": config_key, "active_page": "config"},
    )


@role_required(allowed_roles=["admin"])
def admin_config_edit_view(request, config_key, item_id):
    config = CONFIG_REGISTRY.get(config_key)
    if not config:
        return HttpResponse("Config not found.", status=404)

    model_class = config["model"]
    form_class = config["form"]
    title = config["title"]
    item = get_object_or_404(model_class, id=item_id)

    if request.method == "POST":
        form = form_class(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f"{title} updated.")
        else:
            messages.error(request, f"Failed to update {title}. Please ensure submitted fields are valid.")
            
    return redirect("admin_config_list", config_key=config_key)


@role_required(allowed_roles=["admin"])
def admin_config_delete_view(request, config_key, item_id):
    config = CONFIG_REGISTRY.get(config_key)
    if not config:
        return HttpResponse("Config not found.", status=404)

    model_class = config["model"]
    title = config["title"]
    item = get_object_or_404(model_class, id=item_id)

    if request.method == "POST":
        item.delete()
        messages.success(request, f"{title} deleted.")
        return redirect("admin_config_list", config_key=config_key)

    return render(
        request,
        ADMIN_CONFIG_DELETE_TEMPLATE,
        {"title": title, "config_key": config_key, "item": item, "active_page": "config"},
    )


@role_required(allowed_roles=["admin"])
def manage_mentee_assignments_view(request):
    assignments = MentorMenteeAssignment.objects.select_related("mentor", "mentee").order_by("-start_date")
    if request.method == "POST":
        form = MentorMenteeAssignmentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Assignment saved.")
            return redirect("manage_mentee_assignments")
    else:
        form = MentorMenteeAssignmentForm(initial={"start_date": timezone.now().date()})

    return render(request, ADMIN_ASSIGNMENTS_MENTEE_TEMPLATE, {"assignments": assignments, "form": form, "active_page": "mentee_assign"})


@role_required(allowed_roles=["admin"])
def manage_volunteer_reporting_assignments_view(request):
    assignments = VolunteerReportingAssignment.objects.select_related(
        "volunteer", "programme", "location", "endorser", "assigned_by"
    ).order_by("volunteer__first_name", "volunteer__username")

    if request.method == "POST":
        form = VolunteerReportingAssignmentForm(request.POST)
        if form.is_valid():
            volunteer = form.cleaned_data["volunteer"]
            defaults = {
                "programme": form.cleaned_data["programme"],
                "location": form.cleaned_data["location"],
                "endorser": form.cleaned_data["endorser"],
                "assigned_by": request.user,
            }
            VolunteerReportingAssignment.objects.update_or_create(
                volunteer=volunteer,
                defaults=defaults,
            )
            messages.success(request, "Volunteer reporting assignment saved.")
            return redirect("manage_volunteer_reporting_assignments")
    else:
        form = VolunteerReportingAssignmentForm()

    return render(
        request,
        ADMIN_ASSIGNMENTS_VOLUNTEER_REPORTING_TEMPLATE,
        {"assignments": assignments, "form": form, "active_page": "volunteer_reporting_assign"},
    )


@role_required(allowed_roles=["admin"])
def manage_user_view(request):
    users = User.objects.exclude(role="admin").order_by("id")
    return render(
        request,
        ADMIN_USERS_MANAGE_TEMPLATE,
        {
            "users": users,
            "role_choices": ROLE_CHOICES,
            "active_page": "users",
        },
    )


@role_required(allowed_roles=["admin"])
def user_role_manager_view(request):
    users = CustomUser.objects.order_by("first_name", "last_name")
    role_choices = ROLE_CHOICES
    return render(request, ADMIN_USERS_ROLE_MANAGER_TEMPLATE, {
        "users": users,
        "role_choices": role_choices,
        "active_page": "role_manager",
    })


@role_required(allowed_roles=["admin"])
@require_POST
def update_user_roles_view(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    roles = request.POST.getlist("roles")
    primary_role = request.POST.get("primary_role")

    valid_keys = CustomUser.ROLE_KEYS
    roles = [r for r in roles if r in valid_keys]

    if not roles:
        messages.error(request, f"At least one role is required for {user.get_full_name()}.")
        return redirect("user_role_manager")

    if primary_role and primary_role in roles:
        user.role = primary_role
    else:
        user.role = roles[0]

    user.roles = roles
    user.save()
    messages.success(request, f"Roles updated for {user.get_full_name()}.")
    return redirect("user_role_manager")


@role_required(allowed_roles=["admin"])
def add_user_view(request):
    if request.method != "POST":
        return redirect("manage_user")

    first_name = request.POST.get("first_name")
    last_name = request.POST.get("last_name")
    email = request.POST.get("email")
    phone = request.POST.get("phone")
    role = request.POST.get("role")
    roles = request.POST.getlist("roles")
    current_year = request.POST.get("current_year")
    password = request.POST.get("password")

    if not all([first_name, last_name, email, phone, role]):
        messages.error(request, "All fields are required.")
        return redirect("manage_user")
    if role not in ROLE_KEYS:
        messages.error(request, "Invalid role selected.")
        return redirect("manage_user")

    if User.objects.filter(email=email).exists():
        messages.error(request, "User with this email already exists.")
        return redirect("manage_user")

    # Ensure the primary role is in the roles list
    roles = [item for item in roles if item in ROLE_KEYS]
    if roles:
        if role not in roles:
            roles.insert(0, role)
    else:
        roles = [role]

    user = CustomUser.objects.create(
        username=email,
        email=email,
        first_name=first_name,
        last_name=last_name,
        role=role,
        roles=roles,
        phone=phone,
    )
    if password:
        user.set_password(password)
    else:
        user.set_unusable_password()
    user.save()

    if role == "mentee":
        Mentee.objects.create(user=user, full_name=f"{first_name} {last_name}".strip(), current_year=int(current_year) if current_year else 1)

    messages.success(request, "User added successfully.")
    return redirect("manage_user")


@role_required(allowed_roles=["admin"])
def view_user(request, user_id):
    viewed_user = get_object_or_404(CustomUser, id=user_id)
    activities = Activity.objects.filter(user=viewed_user).order_by("-date")
    return render(request, ADMIN_USERS_DETAIL_TEMPLATE, {"viewed_user": viewed_user, "activities": activities, "active_page": "users"})


@role_required(allowed_roles=["admin"])
def export_user_activities_excel(request, user_id):
    viewed_user = get_object_or_404(CustomUser, id=user_id)
    activities = Activity.objects.filter(user=viewed_user).order_by("-date")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Activities"
    worksheet.append(["Activity", "Duration", "Date", "Feedback", "Learnings", "Other Activity"])

    for activity in activities:
        worksheet.append([
            activity.activity,
            activity.duration,
            activity.date.strftime("%Y-%m-%d") if activity.date else "",
            activity.feedback or "",
            activity.learnings or "",
            activity.other_activity or "",
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{viewed_user.username}_activities.xlsx"'
    workbook.save(response)
    return response


@role_required(allowed_roles=["admin"])
def download_user_template(request):
    response = HttpResponse(content_type="text/csv", headers={"Content-Disposition": 'attachment; filename="user_template.csv"'})
    writer = csv.writer(response)
    writer.writerow(["FIRST_NAME", "LAST_NAME", "EMAIL", "PHONENO", "ROLE"])
    return response


@role_required(allowed_roles=["admin"])
def bulk_upload_users_view(request):
    if request.method == "POST":
        csv_file = request.FILES.get("file")
        if not csv_file:
            messages.error(request, "No file uploaded.")
            return redirect("manage_user")
        if not csv_file.name.endswith(".csv"):
            messages.error(request, "Uploaded file is not a CSV.")
            return redirect("manage_user")

        try:
            decoded_file = csv_file.read().decode("utf-8").splitlines()
            reader = csv.DictReader(decoded_file)
        except Exception:
            messages.error(request, "Error reading the CSV. Ensure it is properly formatted.")
            return redirect("manage_user")

        reader = [{key.strip(): value for key, value in row.items()} for row in reader]
        required_fields = ["FIRST_NAME", "LAST_NAME", "EMAIL", "PHONENO", "ROLE"]
        added_count = 0
        skipped_rows = []

        for index, row in enumerate(reader, start=2):
            row = {key: (value or "").strip() for key, value in row.items()}
            if not all(row.get(field) for field in required_fields):
                skipped_rows.append(f"Row {index}: One or more required fields are empty.")
                continue

            first_name = row["FIRST_NAME"]
            last_name = row["LAST_NAME"]
            email = row["EMAIL"]
            phone = row["PHONENO"]
            raw_roles = [r.strip().lower() for r in row["ROLE"].split(",") if r.strip()]
            valid_keys = CustomUser.ROLE_KEYS

            invalid = [r for r in raw_roles if r not in valid_keys]
            if invalid or not raw_roles:
                skipped_rows.append(
                    f"Row {index}: Invalid role(s) '{', '.join(invalid or ['empty'])}'. Must be one of: {', '.join(valid_keys)}."
                )
                continue
            role = raw_roles[0]
            roles = raw_roles
            if CustomUser.objects.filter(email=email).exists():
                skipped_rows.append(f"Row {index}: Email '{email}' already exists.")
                continue

            username_base = f"{first_name.lower()}.{last_name.lower()}"
            username = username_base
            counter = 1
            while CustomUser.objects.filter(username=username).exists():
                username = f"{username_base}{counter}"
                counter += 1

            user = CustomUser.objects.create_user(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                phone=phone,
                role=role,
                roles=roles,
                password="DefaultPass123",
            )
            if role == "mentee":
                Mentee.objects.create(user=user, full_name=f"{first_name} {last_name}".strip(), current_year=1)
            added_count += 1

        if added_count:
            messages.success(request, f"{added_count} users added successfully.")
        if skipped_rows:
            skipped_msg = "<br>".join(skipped_rows)
            messages.warning(request, mark_safe(f"Some rows were skipped:<br>{skipped_msg}"))

    return redirect("manage_user")


@role_required(allowed_roles=["admin"])
def manage_assignment(request):
    all_endorsers = CustomUser.objects.filter(role="endorser")
    all_mentors = CustomUser.objects.filter(role="mentor")
    unassigned_endorsers = all_endorsers.filter(mentors__isnull=True)
    unassigned_mentors = all_mentors.exclude(assigned_endorsers__isnull=False)
    endorsers_with_counts = [
        {"endorser": endorser, "mentor_count": endorser.mentors.count(), "mentors": endorser.mentors.all()}
        for endorser in all_endorsers
    ]
    context = {
        "endorsers_with_counts": endorsers_with_counts,
        "unassigned_endorsers": unassigned_endorsers,
        "unassigned_mentors": unassigned_mentors,
        "total_endorsers": all_endorsers.count(),
        "total_mentors": all_mentors.count(),
    }
    context["active_page"] = "endorser_assign"
    return render(request, ADMIN_ASSIGNMENTS_MENTOR_TEMPLATE, context)


@role_required(allowed_roles=["admin"])
@require_POST
def assign_mentors(request):
    try:
        endorser_id = request.POST.get("endorser_id")
        mentor_ids = request.POST.getlist("mentor_ids[]")
        if not endorser_id or not mentor_ids:
            return JsonResponse({"status": "failed", "message": "Missing data"})

        endorser = User.objects.get(id=endorser_id)
        mentors = User.objects.filter(id__in=mentor_ids)
        endorser.mentors.add(*mentors)
        return JsonResponse({"status": "success", "message": "Mentors assigned successfully"})
    except Exception as exc:
        print("Assign error:", exc)
        return JsonResponse({"status": "failed", "message": str(exc)})


@role_required(allowed_roles=["admin"])
@require_POST
def unassign_mentor(request):
    endorser_id = request.POST.get("endorser_id")
    mentor_ids = request.POST.getlist("mentor_ids[]")

    try:
        endorser = CustomUser.objects.get(id=endorser_id)
        mentors = CustomUser.objects.filter(id__in=mentor_ids)
        for mentor in mentors:
            endorser.mentors.remove(mentor)
        return JsonResponse({"status": "success"})
    except Exception as exc:
        return JsonResponse({"status": "error", "message": str(exc)})


@role_required(allowed_roles=["admin"])
def get_assigned_mentors(request):
    if request.method != "GET":
        return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)

    endorser_id = request.GET.get("endorser_id")
    if not endorser_id:
        return JsonResponse({"status": "error", "message": "endorser_id missing"}, status=400)

    try:
        endorser = CustomUser.objects.get(id=endorser_id)
        mentors_qs = endorser.mentors.all().values("id", "first_name", "last_name", "email")
        return JsonResponse({"status": "success", "mentors": list(mentors_qs)})
    except CustomUser.DoesNotExist:
        return JsonResponse({"status": "error", "message": f"Endorser with id={endorser_id} not found"}, status=404)
    except Exception as exc:
        return JsonResponse({"status": "error", "message": str(exc), "trace": traceback.format_exc()}, status=500)


@role_required(allowed_roles=["admin"])
@require_POST
def update_assigned_mentors(request):
    endorser_id = request.POST.get("endorser_id")
    mentor_ids = request.POST.getlist("mentor_ids[]")

    try:
        endorser = CustomUser.objects.get(id=endorser_id)
        endorser.mentors.set(mentor_ids)
        return JsonResponse({"status": "success"})
    except Exception as exc:
        return JsonResponse({"status": "error", "message": str(exc)})


# -------------------- Notification Management (FR-30/31) --------------------
@role_required(allowed_roles=["admin"])
def admin_notification_view(request):
    notifications = Notification.objects.order_by("-created_at")[:50]
    if request.method == "POST":
        form = NotificationForm(request.POST)
        if form.is_valid():
            notification = form.save(commit=False)
            notification.created_by = request.user
            notification.save()
            messages.success(request, f"Notification sent to {notification.get_target_group_display()}.")
            return redirect("admin_notifications")
    else:
        form = NotificationForm()
    return render(request, ADMIN_NOTIFICATIONS_TEMPLATE, {"notifications": notifications, "form": form, "active_page": "admin_notifications"})


@role_required(allowed_roles=["admin"])
@require_POST
def admin_notification_delete_view(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id)
    notification.delete()
    messages.success(request, "Notification deleted.")
    return redirect("admin_notifications")


# -------------------- Mentee Bulk Upload (FR-04) --------------------
MENTEE_CSV_COLUMNS = [
    "FULL_NAME", "REGISTER_NO", "GRADE", "CLASS", "SCHOOL",
    "CHAPTER", "LOCATION", "PROGRAMME", "YEAR", "MENTOR_EMAIL",
    "GUARDIAN_NAME", "GUARDIAN_CONTACT",
]


@role_required(allowed_roles=["admin"])
def download_mentee_template(request):
    response = HttpResponse(
        content_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="mentee_upload_template.csv"'},
    )
    writer = csv.writer(response)
    writer.writerow(MENTEE_CSV_COLUMNS)
    writer.writerow([
        "Jane Doe", "REG001", "10", "A", "Sample School",
        "Chapter A", "LOC01", "DIP", "1", "mentor@example.com",
        "John Doe", "9876543210",
    ])
    return response


@role_required(allowed_roles=["admin"])
def bulk_upload_mentees_view(request):
    upload_logs = MenteeUploadLog.objects.order_by("-created_at")[:10]

    if request.method != "POST":
        return render(request, ADMIN_MENTEE_UPLOAD_TEMPLATE, {"upload_logs": upload_logs, "active_page": "import"})

    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        messages.error(request, "No file uploaded.")
        return redirect("bulk_upload_mentees")

    file_name = uploaded_file.name.lower()
    rows = []

    try:
        if file_name.endswith(".csv"):
            decoded = uploaded_file.read().decode("utf-8").splitlines()
            reader = csv.DictReader(decoded)
            rows = [{k.strip(): v for k, v in row.items()} for row in reader]
        elif file_name.endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(uploaded_file, read_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: str(v or "").strip() for i, v in enumerate(row) if i < len(headers)})
        else:
            messages.error(request, "Unsupported file format. Please upload CSV or Excel (.xlsx).")
            return redirect("bulk_upload_mentees")
    except Exception:
        messages.error(request, "Error reading the file. Ensure it is properly formatted.")
        return redirect("bulk_upload_mentees")

    # Cache lookups
    schools_map = {s.name.lower(): s for s in School.objects.all()}
    locations_map = {loc.code.lower(): loc for loc in Location.objects.all()}
    chapters_map = {(ch.name.lower(), ch.school.name.lower()): ch for ch in Chapter.objects.select_related("school").all()}
    programmes_map = {p.name.lower(): p for p in Programme.objects.all()}
    mentors_map = {u.email.lower(): u for u in CustomUser.objects.filter(role="mentor")}

    added_count = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        row = {k: (v or "").strip() for k, v in row.items()}
        full_name = row.get("FULL_NAME", "")
        if not full_name:
            errors.append(f"Row {idx}: FULL_NAME is required.")
            continue

        year_str = row.get("YEAR", "1")
        try:
            current_year = int(year_str)
            if current_year not in (1, 2, 3, 4):
                raise ValueError
        except (ValueError, TypeError):
            errors.append(f"Row {idx}: Invalid YEAR '{year_str}'. Must be 1-4.")
            continue

        register_no = row.get("REGISTER_NO", "")
        if register_no and Mentee.objects.filter(register_no=register_no).exists():
            errors.append(f"Row {idx}: Duplicate register number '{register_no}'.")
            continue

        school_name = row.get("SCHOOL", "")
        school = schools_map.get(school_name.lower()) if school_name else None
        if school_name and not school:
            school = School.objects.create(name=school_name)
            schools_map[school_name.lower()] = school

        location_code = row.get("LOCATION", "")
        location = locations_map.get(location_code.lower()) if location_code else None

        chapter_name = row.get("CHAPTER", "")
        chapter = None
        if chapter_name and school:
            chapter = chapters_map.get((chapter_name.lower(), school.name.lower()))

        programme_name = row.get("PROGRAMME", "")
        programme = programmes_map.get(programme_name.lower()) if programme_name else None

        mentor_email = row.get("MENTOR_EMAIL", "")
        mentor = mentors_map.get(mentor_email.lower()) if mentor_email else None
        if mentor_email and not mentor:
            errors.append(f"Row {idx}: Mentor email '{mentor_email}' not found.")
            continue

        Mentee.objects.create(
            full_name=full_name,
            register_no=register_no,
            grade=row.get("GRADE", ""),
            cls=row.get("CLASS", ""),
            school=school,
            chapter=chapter,
            location=location,
            programme_fk=programme,
            programme=programme_name,
            current_year=current_year,
            assigned_mentor=mentor,
            guardian_name=row.get("GUARDIAN_NAME", ""),
            guardian_contact=row.get("GUARDIAN_CONTACT", ""),
        )
        added_count += 1

    # Log upload
    MenteeUploadLog.objects.create(
        file_name=uploaded_file.name,
        uploaded_by=request.user,
        total_rows=len(rows),
        success_count=added_count,
        error_count=len(errors),
        error_details=errors,
    )

    if added_count:
        messages.success(request, f"{added_count} mentees imported successfully.")
    if errors:
        display_errors = errors[:20]
        msg = "<br>".join(display_errors)
        if len(errors) > 20:
            msg += f"<br>... and {len(errors) - 20} more errors."
        messages.warning(request, mark_safe(f"Some rows had issues:<br>{msg}"))

    return redirect("bulk_upload_mentees")


@role_required(allowed_roles=["admin"])
def mentee_upload_log_view(request):
    logs = MenteeUploadLog.objects.order_by("-created_at")[:50]
    return render(request, ADMIN_MENTEE_UPLOAD_LOG_TEMPLATE, {"logs": logs, "active_page": "upload_logs"})


@role_required(allowed_roles=["admin"])
def export_reflective_reports_excel(request):
    reports = ReflectiveReport.objects.select_related("user", "programme", "location", "endorser").all().order_by("-date")
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Reflective Reports"
    worksheet.append(["User", "Role", "Programme", "Location", "Activity Name", "Duration", "Date", "Endorser", "Learnings", "Feedback", "Status"])
    
    for report in reports:
        worksheet.append([
            report.user.get_full_name() or report.user.username,
            report.user.get_roles_display()[0] if report.user.roles else report.user.role,
            report.programme.name if report.programme else "",
            report.location.name if report.location else "",
            report.activity_name,
            float(report.duration) if report.duration else 0.0,
            report.date.strftime("%Y-%m-%d") if report.date else "",
            report.endorser.get_full_name() if report.endorser else "",
            report.learnings or "",
            report.feedback or "",
            report.get_status_display()
        ])
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="reflective_reports_export.xlsx"'
    workbook.save(response)
    return response


@role_required(allowed_roles=["admin"])
def export_work_diaries_excel(request):
    entries = DiaryEntry.objects.select_related("volunteer", "location").all().order_by("-date")
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Work Diaries"
    worksheet.append(["Volunteer", "Date", "Duration", "Location", "Linked Activity", "Narrative Entry", "Review Status"])
    
    for entry in entries:
        worksheet.append([
            entry.volunteer.get_full_name() or entry.volunteer.username,
            entry.date.strftime("%Y-%m-%d") if entry.date else "",
            float(entry.duration) if entry.duration else 0.0,
            entry.location.name if entry.location else "",
            entry.linked_activity or "",
            entry.narrative_entry or "",
            entry.get_review_status_display()
        ])
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="work_diaries_export.xlsx"'
    workbook.save(response)
    return response
